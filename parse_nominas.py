"""Parser unificado de nóminas CRUCH 2022-2026.

Output: data/nominas/{año}.parquet con schema estandarizado:
    rut, dv, rut_dv, apellido_paterno, apellido_materno, nombres,
    monto_utm, universidad, year
"""
import sys
import openpyxl
import pdfplumber
import pandas as pd
import re
from pathlib import Path

RAW = Path("/Users/antonio/deudores-fscu/data/nominas_raw")
OUT = Path("/Users/antonio/deudores-fscu/data/nominas")
OUT.mkdir(parents=True, exist_ok=True)

DV_RE = re.compile(r"^[0-9kK]$")
NUM_RE = re.compile(r"^\d+$")

def clean_dv(x):
    s = str(x).strip().upper()
    return s if DV_RE.match(s) else None

def clean_rut(x):
    s = str(x).strip().lstrip('0')
    return s if NUM_RE.match(s) else None

def parse_xlsx(year, path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().upper().replace(' ','_') if h else '' for h in next(rows)]
    out = []
    for r in rows:
        d = dict(zip(header, r))
        rut = clean_rut(d.get('RUT'))
        dv = clean_dv(d.get('DIG_VERIF'))
        monto = d.get('MONTO_MOROSO_UTM')
        if not (rut and dv and monto is not None): continue
        try: monto_f = float(monto)
        except (ValueError, TypeError): continue
        # En 2022 "APELLIDO_PATERNO" aparece 2 veces (el segundo es materno);
        # openpyxl lo colapsa, así que tomamos el valor tal cual.
        # Para 2023 hay columnas diferenciadas.
        apat = d.get('APELLIDO_PATERNO') or ''
        amat = d.get('APELLIDO_MATERNO') or ''
        # si amat vacío y es 2022, la columna es la 2da "APELLIDO_PATERNO"
        # pero openpyxl la sobreescribió. Lo inferimos releyendo posiciones:
        out.append({
            'rut': rut, 'dv': dv,
            'apellido_paterno': str(apat).strip(),
            'apellido_materno': str(amat).strip(),
            'nombres': str(d.get('NOMBRE') or d.get('NOMBRES') or '').strip(),
            'monto_utm': monto_f,
            'universidad': str(d.get('UNIVERSIDAD') or '').strip(),
            'year': year,
        })
    return pd.DataFrame(out)

def parse_xlsx_positional(year, path):
    """Parser posicional — evita colisiones de header."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    ncols = len(rows[0])
    # 2022: 7 cols [RUT, DV, ApPat, ApMat, Nombre, Monto, Universidad]
    # 2023: 8 cols [RUT, DV, ApPat, ApMat, Nombres, Monto, Cod_Univ, Universidad]
    idx_rut = 0; idx_dv = 1; idx_apat = 2; idx_amat = 3; idx_nom = 4; idx_monto = 5
    idx_univ = 6 if ncols == 7 else 7
    out = []
    for r in rows[1:]:
        rut = clean_rut(r[idx_rut]); dv = clean_dv(r[idx_dv])
        monto = r[idx_monto]
        if not (rut and dv and monto is not None): continue
        try: monto_f = float(monto)
        except (ValueError, TypeError): continue
        out.append({
            'rut': rut, 'dv': dv,
            'apellido_paterno': str(r[idx_apat] or '').strip(),
            'apellido_materno': str(r[idx_amat] or '').strip(),
            'nombres': str(r[idx_nom] or '').strip(),
            'monto_utm': monto_f,
            'universidad': str(r[idx_univ] or '').strip(),
            'year': year,
        })
    return pd.DataFrame(out)

PDF_TABLE_RE = re.compile(
    # row_num rut dv ap_pat ap_mat nombres monto 1 universidad
    r"^\s*\d+\s+(\d+)\s+([0-9kK])\s+(\S+)\s+(\S+)\s+(.+?)\s+(\d+(?:[.,]\d+)?)\s+\d+\s+(\S.+?)\s*$"
)
PDF_TEXT_RE = re.compile(
    # rut (7-9 digits), dv (0-9 o K), lookahead para anclar nombre en letra mayúscula
    # (evita que el regex interprete 1 dígito del nombre como parte del RUT)
    r"^\s*(\d{7,9})\s*([0-9kK])\s*(?=[A-ZÁÉÍÓÚÑ])(.+?)\s+(\d+(?:[.,]\d+)?)\s+\d{1,2}([A-ZÁÉÍÓÚÑ].+?)\s*$"
)

def parse_pdf(year, path, progress_every=100):
    """Detecta auto-modo: tabla (2026) vs texto plano (2024-2025)."""
    pdf = pdfplumber.open(path)
    n = len(pdf.pages)
    # Decide modo con página a mitad
    mid = pdf.pages[min(50, n-1)]
    mode = 'table' if len(mid.extract_tables()) > 0 and any(r for r in mid.extract_tables()[0] if r and len(r) >= 9) else 'text'
    print(f"[{year}] PDF {n} págs · modo={mode}", flush=True)
    rows = []
    for i, page in enumerate(pdf.pages):
        if mode == 'table':
            for tbl in page.extract_tables():
                for r in tbl:
                    if not r or len(r) < 9: continue
                    rut = clean_rut(r[1]); dv = clean_dv(r[2])
                    if not (rut and dv): continue
                    try: monto_f = float(str(r[6]).replace(',','.'))
                    except (ValueError, TypeError): continue
                    rows.append({
                        'rut': rut, 'dv': dv,
                        'apellido_paterno': str(r[3] or '').strip(),
                        'apellido_materno': str(r[4] or '').strip(),
                        'nombres': str(r[5] or '').strip(),
                        'monto_utm': monto_f,
                        'universidad': str(r[8] or '').strip(),
                        'year': year,
                    })
        else:
            txt = page.extract_text() or ''
            for ln in txt.splitlines():
                m = PDF_TEXT_RE.match(ln)
                if not m: continue
                rut, dv, nombre_full, monto, univ = m.groups()
                try: monto_f = float(monto.replace(',','.'))
                except ValueError: continue
                parts = nombre_full.split()
                ap_p = parts[0] if len(parts) >= 1 else ''
                ap_m = parts[1] if len(parts) >= 2 else ''
                nombres = ' '.join(parts[2:]) if len(parts) >= 3 else ''
                rows.append({
                    'rut': clean_rut(rut), 'dv': clean_dv(dv),
                    'apellido_paterno': ap_p, 'apellido_materno': ap_m,
                    'nombres': nombres, 'monto_utm': monto_f,
                    'universidad': univ.strip(), 'year': year,
                })
        if (i+1) % progress_every == 0 or i == n-1:
            print(f"[{year}]  {i+1}/{n}  rows={len(rows):,}", flush=True)
    pdf.close()
    return pd.DataFrame(rows)

def run(year, typ):
    src = RAW / f"{year}.{typ}"
    dst = OUT / f"{year}.parquet"
    if typ == "xlsx":
        df = parse_xlsx_positional(year, src)
    else:
        df = parse_pdf(year, src)
    df['rut_dv'] = df['rut'] + '-' + df['dv']
    df.to_parquet(dst, index=False)
    print(f"\n[{year}] Escrito {dst}")
    print(f"[{year}]  filas: {len(df):,}")
    print(f"[{year}]  RUTs únicos: {df['rut'].nunique():,}")
    print(f"[{year}]  Monto total UTM: {df['monto_utm'].sum():,.0f}")
    print(f"[{year}]  Universidades: {df['universidad'].nunique()}")

if __name__ == "__main__":
    year = int(sys.argv[1])
    typ = sys.argv[2]
    run(year, typ)
